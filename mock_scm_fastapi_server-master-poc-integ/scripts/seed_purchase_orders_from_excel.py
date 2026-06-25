import argparse
import json
import re
import uuid
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise ImportError(
        "openpyxl is required to run this script. Install it with `pip install openpyxl`."
    ) from exc

BASE_DIR = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_FILE = BASE_DIR / "data" / "purchase_orders.json"
DEFAULT_CANONICAL_DIR = BASE_DIR / "data" / "canonical"

PO_FIELD_CANDIDATES = [
    "po_number",
    "po_no",
    "po",
    "purchase_order",
    "purchase_order_number",
]

SUPPLIER_FIELD_CANDIDATES = [
    "local_supplier_nm",
    "supplier_name",
    "supplier",
    "vendor",
    "vendor_name",
]

SUPPLIER_ID_FIELD_CANDIDATES = [
    "local_supplier_id",
    "msid",
    "src_msid",
    "supplier_id",
    "vendor_id",
    "suppliercode",
    "vendor_code",
]

SUPPLIER_EMAIL_FIELD_CANDIDATES = ["supplier_email", "email", "vendor_email"]

SITE_FIELD_CANDIDATES = ["site_cd", "site", "location", "plant", "vendor_site_id", "location_id"]

STATUS_FIELD_CANDIDATES = ["status", "po_status", "purchase_order_status"]

SOURCE_FIELD_CANDIDATES = ["source_system_cd", "source_system", "source", "system"]

PERIOD_DATE_CANDIDATES = ["period_dt", "period_date"]

PO_RELEASE_CANDIDATES = ["po_release_no", "release_no"]

PO_LINE_REVISION_CANDIDATES = ["po_line_revision_no", "line_revision_no"]

PO_LINE_ISSUE_DATE_CANDIDATES = ["po_line_issue_dt", "po_line_issue_date"]

ORIGINAL_PROMISE_CANDIDATES = [
    "original_prom_date",
    "original_promise_date",
    "original_promise_date1",
    "original_promise_date2",
    "original_promise_date3",
    "original_promise_date4",
]

OTS_PROMISE_CANDIDATES = ["ots_promise_date", "ots_promise__date", "shipment_date"]

PO_LINE_ACK_STATUS_CANDIDATES = ["po_line_ackn_status", "po_line_ack_status"]

PO_LINE_ACK_DATE_CANDIDATES = ["po_line_ackn_dt", "po_line_ack_date"]

SHIPMENT_MODE_CANDIDATES = ["shipment_mode", "ship_via_carrier", "transportation"]

SAVINGS_TYPE_CANDIDATES = ["savings_type"]

SAVINGS_CANDIDATES = ["savings"]

STD_UNIT_COST_CANDIDATES = ["std_unit_cost"]

ERP_EXTRACT_CANDIDATES = ["erp_extract_date"]

EXCEPTION_CANDIDATES = ["except_message", "mrp_exceptions"]

RESCHEDULING_CANDIDATES = ["rescheduling_date"]

PO_FEEDBACK_CANDIDATES = ["po_feedback"]

PURCHASING_GROUP_CANDIDATES = ["purchasing_group"]

INCOTERM_CANDIDATES = ["incoterm"]

INCOTERM_PLACE_CANDIDATES = ["incoterm_named_place"]

ITEM_CATEGORY_CANDIDATES = ["item_category_id", "category_cd_site"]

QUANTITY_OUTSTANDING_CANDIDATES = ["quantity_outstanding"]

CURRENCY_CANDIDATES = ["currency", "currency_code"]

DELIVERY_FIELD_CANDIDATES = [
    "latest_promise_date",
    "ots_promise_date",
    "delivery_date",
    "delivery",
    "ship_date",
]

MRP_FIELD_CANDIDATES = ["mrp_need_by_dt", "mrp_need_by_date", "mrp_date", "mrp", "mrp_exceptions"]

CREATED_DATE_CANDIDATES = ["po_issue_dt", "created_date", "creation_date", "po_date"]

PAYMENT_TERMS_CANDIDATES = ["payment_term", "payment_terms", "terms"]

PROCUREMENT_SPECIALIST_CANDIDATES = ["procurement_specialist_id", "po_created_by", "ps_id", "buyer_id"]

TOTAL_VALUE_CANDIDATES = ["total_value", "po_value", "amount", "total_amount"]

LINE_NUMBER_CANDIDATES = ["po_line_no", "line_number", "line_no", "line"]

MATERIAL_CODE_CANDIDATES = [
    "item_no",
    "material_code",
    "item_code",
    "part_number",
    "material",
    "part_no",
]

DESCRIPTION_CANDIDATES = [
    "item_desc",
    "description",
    "item_description",
    "part_description",
    "material_description",
]

QUANTITY_CANDIDATES = ["quantity_ord", "quantity", "qty", "order_qty", "ordered_quantity"]

UNIT_PRICE_CANDIDATES = ["unit_cost", "unit_price", "price", "amount"]

LINE_TOTAL_CANDIDATES = ["line_total", "ext_price", "extended_price", "amount"]

DEFAULT_STATUS = "CREATED"
DEFAULT_SOURCE = "SAP"
DEFAULT_CURRENCY = "USD"
DEFAULT_PAYMENT_TERMS = "Net 30"
DEFAULT_MRP_EXCEPTIONS = "NONE"
DEFAULT_PROCSPEC_ID = "PS-001"

NULL_STRINGS = {"NULL", "NONE", "N/A", "NA", "NAN"}


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    normalized = re.sub(r"[^0-9a-z]+", "_", str(value).strip().lower())
    return normalized.strip("_")


def deserialize_date(value: Any) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    if text.upper() in NULL_STRINGS:
        return None
    return text


def deserialize_number(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if text.upper() in NULL_STRINGS:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_value(value: Any) -> Optional[Any]:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        if text.upper() in NULL_STRINGS:
            return None
        return text
    return value


def normalize_text_like(value: Any) -> Optional[str]:
    normalized = normalize_value(value)
    if normalized is None:
        return None
    if isinstance(normalized, date):
        return normalized.isoformat()
    return str(normalized).strip() or None


def find_first_value(row: Dict[str, Any], candidates: List[str]) -> Optional[Any]:
    for key in candidates:
        if key in row:
            normalized = normalize_value(row[key])
            if normalized is not None:
                return normalized
    return None


def build_supplier_id(name: Optional[str], supplier_map: Dict[str, str], supplier_counter: List[int]) -> str:
    if name:
        key = normalize_header(name)
        if key in supplier_map:
            return supplier_map[key]
        supplier_id = f"SUP-{supplier_counter[0]:03d}"
        supplier_map[key] = supplier_id
        supplier_counter[0] += 1
        return supplier_id

    supplier_id = f"SUP-{supplier_counter[0]:03d}"
    supplier_counter[0] += 1
    return supplier_id


def build_supplier_email(name: Optional[str], supplier_id: str) -> str:
    if isinstance(name, str) and "@" in name:
        return name.strip()
    if isinstance(name, str):
        candidate = re.sub(r"[^0-9a-z]+", "", name.strip().lower())
        if candidate:
            return f"{candidate}@mockscm.com"
    return f"{supplier_id.lower()}@mockscm.com"


def load_rows_from_excel(path: Path, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel sheet is empty")

    headers = [normalize_header(cell) for cell in rows[0]]
    data_rows: List[Dict[str, Any]] = []
    for row in rows[1:]:
        if row is None:
            continue
        data = {headers[idx]: cell for idx, cell in enumerate(row) if idx < len(headers)}
        if not any(value not in (None, "") for value in data.values()):
            continue
        data_rows.append(data)
    return data_rows


def group_rows_by_po(data_rows: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in data_rows:
        po_number = find_first_value(row, PO_FIELD_CANDIDATES)
        if not po_number:
            raise ValueError("Missing PO number in one or more rows")
        groups[str(po_number).strip()].append(row)
    return groups


def make_line_item(row: Dict[str, Any], line_index: int) -> Dict[str, Any]:
    line_number_raw = find_first_value(row, LINE_NUMBER_CANDIDATES)
    if line_number_raw is None or line_number_raw == "":
        line_number: Any = line_index
    else:
        line_number = str(line_number_raw).strip()

    item_no = find_first_value(row, ["item_no"])
    material_code = find_first_value(row, MATERIAL_CODE_CANDIDATES) or f"MAT-{line_index:04d}"
    description = find_first_value(row, DESCRIPTION_CANDIDATES) or "Material"
    quantity = deserialize_number(find_first_value(row, QUANTITY_CANDIDATES)) or 1.0
    unit_price = deserialize_number(find_first_value(row, UNIT_PRICE_CANDIDATES))
    line_total = deserialize_number(find_first_value(row, LINE_TOTAL_CANDIDATES))

    if unit_price is None and line_total is not None and quantity:
        unit_price = line_total / max(quantity, 1.0)
    if unit_price is None:
        unit_price = 0.0

    return {
        "line_number": line_number,
        "po_line_no": line_number,
        "po_release_no": deserialize_number(find_first_value(row, PO_RELEASE_CANDIDATES)),
        "po_line_revision_no": deserialize_number(find_first_value(row, PO_LINE_REVISION_CANDIDATES)),
        "po_line_issue_date": deserialize_date(find_first_value(row, PO_LINE_ISSUE_DATE_CANDIDATES)),
        "item_no": str(item_no).strip() if item_no not in (None, "") else str(material_code).strip(),
        "material_code": str(material_code).strip(),
        "description": str(description).strip(),
        "quantity": int(quantity) if float(quantity).is_integer() else float(quantity),
        "quantity_outstanding": deserialize_number(find_first_value(row, QUANTITY_OUTSTANDING_CANDIDATES)),
        "unit_price": round(float(unit_price), 2),
        "currency_code": find_first_value(row, CURRENCY_CANDIDATES),
        "mrp_need_by_date": deserialize_date(find_first_value(row, MRP_FIELD_CANDIDATES)),
        "original_promise_date": deserialize_date(find_first_value(row, ORIGINAL_PROMISE_CANDIDATES)),
        "latest_promise_date": deserialize_date(find_first_value(row, DELIVERY_FIELD_CANDIDATES)),
        "ots_promise_date": deserialize_date(find_first_value(row, OTS_PROMISE_CANDIDATES)),
        "item_category_id": find_first_value(row, ITEM_CATEGORY_CANDIDATES),
        "incoterm": find_first_value(row, INCOTERM_CANDIDATES),
        "incoterm_named_place": find_first_value(row, INCOTERM_PLACE_CANDIDATES),
        "shipment_mode": find_first_value(row, SHIPMENT_MODE_CANDIDATES),
        "po_line_ack_status": find_first_value(row, PO_LINE_ACK_STATUS_CANDIDATES),
        "po_line_ack_date": deserialize_date(find_first_value(row, PO_LINE_ACK_DATE_CANDIDATES)),
        "savings_type": find_first_value(row, SAVINGS_TYPE_CANDIDATES),
        "savings": deserialize_number(find_first_value(row, SAVINGS_CANDIDATES)),
        "std_unit_cost": deserialize_number(find_first_value(row, STD_UNIT_COST_CANDIDATES)),
        "erp_extract_date": deserialize_date(find_first_value(row, ERP_EXTRACT_CANDIDATES)),
        "except_message": normalize_text_like(find_first_value(row, EXCEPTION_CANDIDATES)),
        "rescheduling_date": deserialize_date(find_first_value(row, RESCHEDULING_CANDIDATES)),
        "po_feedback": normalize_text_like(find_first_value(row, PO_FEEDBACK_CANDIDATES)),
        "drawing_no": normalize_text_like(find_first_value(row, ["drawing_no"])),
        "drawing_revision": normalize_text_like(find_first_value(row, ["drawing_rev", "drawing_revision"])),
        "seals_ord_no": normalize_text_like(find_first_value(row, ["sales_ord_no", "seals_ord_no"])),
        "purchasing_group": normalize_text_like(find_first_value(row, PURCHASING_GROUP_CANDIDATES)),
    }


def merge_po_fields(rows: List[Dict[str, Any]], supplier_map: Dict[str, str], supplier_counter: List[int]) -> Dict[str, Any]:
    first_row = rows[0]

    supplier_name = find_first_value(first_row, SUPPLIER_FIELD_CANDIDATES)
    supplier_id = find_first_value(first_row, SUPPLIER_ID_FIELD_CANDIDATES)
    if not supplier_id:
        supplier_id = build_supplier_id(str(supplier_name) if supplier_name else None, supplier_map, supplier_counter)
    supplier_email = find_first_value(first_row, SUPPLIER_EMAIL_FIELD_CANDIDATES) or build_supplier_email(
        str(supplier_name) if supplier_name else None, supplier_id
    )
    supplier_id_text = str(supplier_id).strip()
    supplier_digits = "".join(ch for ch in supplier_id_text if ch.isdigit())
    supplier_suffix = supplier_digits[-3:] if supplier_digits else supplier_id_text[-3:]

    site = find_first_value(first_row, SITE_FIELD_CANDIDATES) or f"Site-{supplier_suffix}"
    location_id = find_first_value(first_row, ["location_id"])

    po_number = str(find_first_value(first_row, PO_FIELD_CANDIDATES)).strip()
    status = str(find_first_value(first_row, STATUS_FIELD_CANDIDATES) or DEFAULT_STATUS).strip().upper()
    source_system = str(find_first_value(first_row, SOURCE_FIELD_CANDIDATES) or DEFAULT_SOURCE).strip()
    period_date = deserialize_date(find_first_value(first_row, PERIOD_DATE_CANDIDATES))
    delivery_date = deserialize_date(find_first_value(first_row, DELIVERY_FIELD_CANDIDATES))
    mrp_need_by_date = deserialize_date(find_first_value(first_row, MRP_FIELD_CANDIDATES))
    po_issue_date = deserialize_date(find_first_value(first_row, ["po_issue_dt", "po_issue_date"]))
    created_date = deserialize_date(find_first_value(first_row, CREATED_DATE_CANDIDATES)) or date.today().isoformat()
    payment_terms = str(find_first_value(first_row, PAYMENT_TERMS_CANDIDATES) or DEFAULT_PAYMENT_TERMS).strip()
    procurement_specialist_id = str(find_first_value(first_row, PROCUREMENT_SPECIALIST_CANDIDATES) or DEFAULT_PROCSPEC_ID).strip()
    total_value = deserialize_number(find_first_value(first_row, TOTAL_VALUE_CANDIDATES))
    purchasing_group = normalize_text_like(find_first_value(first_row, PURCHASING_GROUP_CANDIDATES))
    except_message = normalize_text_like(find_first_value(first_row, EXCEPTION_CANDIDATES))
    incoterm = normalize_text_like(find_first_value(first_row, INCOTERM_CANDIDATES))
    incoterm_named_place = normalize_text_like(find_first_value(first_row, INCOTERM_PLACE_CANDIDATES))

    line_items: List[Dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        material = find_first_value(row, MATERIAL_CODE_CANDIDATES)
        quantity = find_first_value(row, QUANTITY_CANDIDATES)
        unit_price = find_first_value(row, UNIT_PRICE_CANDIDATES)
        if material or quantity or unit_price or find_first_value(row, LINE_TOTAL_CANDIDATES):
            line_items.append(make_line_item(row, index))

    if not line_items:
        line_items.append(
            {
                "line_number": 1,
                "material_code": f"MAT-{po_number[-4:]}",
                "description": "Generated line item",
                "quantity": 1,
                "unit_price": float(total_value or 0.0),
            }
        )

    computed_total = round(sum(item["quantity"] * item["unit_price"] for item in line_items), 2)
    if total_value is None:
        total_value = computed_total

    return {
        "id": str(uuid.uuid4()),
        "po_header_id": po_number,
        "period_date": period_date,
        "po_number": po_number,
        "source_system": source_system,
        "status": status,
        "supplier_id": supplier_id_text,
        "supplier_msid": _supplier_msid_from_supplier_id(supplier_id_text, supplier_counter[0]),
        "local_supplier_id": _supplier_msid_from_supplier_id(supplier_id_text, supplier_counter[0]),
        "supplier_name": str(supplier_name or f"Supplier {supplier_suffix}").strip(),
        "supplier_email": supplier_email,
        "site": str(site),
        "location_id": location_id,
        "procurement_specialist_id": procurement_specialist_id,
        "delegated_user_id": "",
        "currency": DEFAULT_CURRENCY,
        "total_value": float(total_value),
        "po_issue_date": po_issue_date or created_date,
        "delivery_date": delivery_date or date.today().isoformat(),
        "mrp_need_by_date": mrp_need_by_date,
        "payment_terms": payment_terms,
        "purchasing_group": purchasing_group,
        "incoterm": incoterm,
        "incoterm_named_place": incoterm_named_place,
        "mrp_exceptions": except_message or DEFAULT_MRP_EXCEPTIONS,
        "last_modified_by": procurement_specialist_id,
        "last_modified_date": delivery_date or created_date,
        "created_date": created_date,
        "revision_changes": 0,
        "line_items": line_items,
    }


def _supplier_msid_from_supplier_id(supplier_id: str, fallback: int) -> int:
    digits = "".join(ch for ch in supplier_id if ch.isdigit())
    if digits:
        return int(digits)
    return fallback


def _location_id_from_site(site: str, fallback: int) -> int:
    digits = "".join(ch for ch in site if ch.isdigit())
    if digits:
        return int(digits)
    return fallback


def _derive_reference_data(purchase_orders: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    suppliers_map: Dict[int, Dict[str, Any]] = {}
    locations_map: Dict[int, Dict[str, Any]] = {}
    items_map: Dict[str, Dict[str, Any]] = {}

    supplier_fallback = 1000
    location_fallback = 2000

    for order in purchase_orders:
        supplier_id = str(order.get("supplier_id") or "")
        supplier_name = str(order.get("supplier_name") or "").strip()
        supplier_email = str(order.get("supplier_email") or "").strip()
        supplier_msid = _supplier_msid_from_supplier_id(supplier_id, supplier_fallback)
        supplier_fallback = max(supplier_fallback, supplier_msid + 1)

        if supplier_msid not in suppliers_map:
            suppliers_map[supplier_msid] = {
                "msid": supplier_msid,
                "supplier_name": supplier_name or f"Supplier {supplier_msid}",
                "supplier_dba_name": None,
                "category_id": None,
                "category_id2": None,
                "slp_id": None,
                "address": "",
                "city": "",
                "state_province": "",
                "iso_country_code": "US",
                "postal_code": "",
                "payment_term": str(order.get("payment_terms") or DEFAULT_PAYMENT_TERMS),
                "incoterm": None,
                "segmentation": None,
                "tactical_approach": None,
                "approval_status": None,
                "scobc_ack": None,
                "slp_nda_ack": None,
                "scobc_received": None,
                "scobc_understood": None,
                "company_size": None,
                "scobc_accept": None,
                "is_parent": None,
                "duns_no": None,
                "bp_type": None,
                "mdg_managed": None,
                "bp_block": None,
                "posting_block": None,
                "po_block": None,
                "diversity": None,
                "management_model": None,
                "assigned_sqe": None,
                "supplier_manager": None,
                "due_diligence": None,
                "is_archived": False,
                "supplier_business_focus": None,
                "seed_email": supplier_email,
                "seed_user_id": supplier_id,
            }

        site = str(order.get("site") or "").strip() or f"Site-{location_fallback}"
        location_id_raw = order.get("location_id")
        if location_id_raw in (None, ""):
            location_id = _location_id_from_site(site, location_fallback)
        else:
            location_id = _location_id_from_site(str(location_id_raw), location_fallback)
        location_fallback = max(location_fallback, location_id + 1)

        if location_id not in locations_map:
            locations_map[location_id] = {
                "location_id": location_id,
                "location_name": site,
                "platform": "UNKNOWN",
                "iso_country_code": "US",
                "address": "",
                "city": "",
                "state_province": "",
                "postal_code": "",
                "operation": "",
                "sector": "",
                "division": "",
                "istp_flag": None,
                "location_status": True,
                "location_type": "",
                "heritage_name": "",
                "operating_model": "",
                "platform_management_region": "",
                "is_balanced_scorecard": False,
                "business_unit": "",
                "ru_no": "",
                "is_archived": False,
                "custom_bu": "",
            }

        for line in order.get("line_items", []):
            item_no = str(line.get("material_code") or line.get("item_no") or "").strip()
            if not item_no:
                continue
            if item_no not in items_map:
                items_map[item_no] = {
                    "item_no": item_no,
                    "location_id": location_id,
                    "site_code": site,
                    "item_lead_time": 0,
                    "pattern_no": None,
                    "material_code": item_no,
                    "item_weight": None,
                    "item_weight_unit": "KG",
                    "is_active": True,
                    "is_safety_stock": False,
                    "safety_stock_min": None,
                    "safety_stock_max": None,
                    "stock_level": None,
                }

    suppliers = sorted(suppliers_map.values(), key=lambda row: row["msid"])
    locations = sorted(locations_map.values(), key=lambda row: row["location_id"])
    items = sorted(items_map.values(), key=lambda row: row["item_no"])

    return {
        "suppliers": suppliers,
        "locations": locations,
        "items": items,
    }


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as output_file:
        json.dump(payload, output_file, indent=2)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert Open PO Data Excel file into purchase order JSON seed data"
    )
    parser.add_argument("input", type=Path, help="Path to the Excel workbook")
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Worksheet name to read (default: first sheet)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_FILE,
        help="Output JSON file path",
    )
    parser.add_argument(
        "--canonical-dir",
        type=Path,
        default=DEFAULT_CANONICAL_DIR,
        help="Directory for canonical JSON artifacts",
    )
    parser.add_argument(
        "--canonical-only",
        action="store_true",
        help="Write canonical artifacts only (purchase_orders + suppliers + locations + items)",
    )
    parser.add_argument(
        "--write-derived-masters",
        action="store_true",
        help="Also write derived suppliers/locations/items into data/*.json",
    )
    args = parser.parse_args()

    data_rows = load_rows_from_excel(args.input, args.sheet)
    po_groups = group_rows_by_po(data_rows)

    supplier_map: Dict[str, str] = {}
    supplier_counter = [1]
    purchase_orders: List[Dict[str, Any]] = []

    for rows in po_groups.values():
        purchase_orders.append(merge_po_fields(rows, supplier_map, supplier_counter))

    derived = _derive_reference_data(purchase_orders)

    canonical_po = args.canonical_dir / "purchase_orders.canonical.json"
    canonical_suppliers = args.canonical_dir / "suppliers.canonical.json"
    canonical_locations = args.canonical_dir / "locations.canonical.json"
    canonical_items = args.canonical_dir / "items.canonical.json"

    _write_json(canonical_po, purchase_orders)
    _write_json(canonical_suppliers, derived["suppliers"])
    _write_json(canonical_locations, derived["locations"])
    _write_json(canonical_items, derived["items"])

    if not args.canonical_only:
        _write_json(args.output, purchase_orders)

    if args.write_derived_masters:
        _write_json(BASE_DIR / "data" / "suppliers.json", derived["suppliers"])
        _write_json(BASE_DIR / "data" / "locations.json", derived["locations"])
        _write_json(BASE_DIR / "data" / "items.json", derived["items"])

    print(f"Wrote {len(purchase_orders)} purchase orders to {canonical_po}")
    print(f"Wrote {len(derived['suppliers'])} suppliers to {canonical_suppliers}")
    print(f"Wrote {len(derived['locations'])} locations to {canonical_locations}")
    print(f"Wrote {len(derived['items'])} items to {canonical_items}")
    if not args.canonical_only:
        print(f"Wrote runtime purchase orders to {args.output}")


if __name__ == "__main__":
    main()
